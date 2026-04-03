#!/usr/bin/env python3
"""
持仓导入工具 — 支持从Excel表格和API两种数据源更新portfolio-config.json
用法:
  python import_portfolio.py                     # 从默认Excel模板导入
  python import_portfolio.py /path/to/file.xlsx  # 从指定Excel导入
  python import_portfolio.py --api <url>         # 从API接口导入
  python import_portfolio.py --merge             # 合并模式(Excel+API)
"""

import json
import os
import sys
from datetime import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
CONFIG_PATH = os.path.join(SKILL_DIR, "references", "portfolio-config.json")
DEFAULT_XLSX = os.path.join(SKILL_DIR, "持仓录入模板.xlsx")

# ============================================================
# Excel导入
# ============================================================

def import_from_excel(xlsx_path):
    """从Excel模板导入持仓和关注列表"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装openpyxl: pip install openpyxl --break-system-packages")
        sys.exit(1)

    if not os.path.exists(xlsx_path):
        print(f"❌ 文件不存在: {xlsx_path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ---- 读取持仓 ----
    holdings = []
    if "持仓录入" in wb.sheetnames:
        ws = wb["持仓录入"]
        for row in ws.iter_rows(min_row=4, max_row=50, values_only=True):
            symbol = row[0]
            if not symbol or str(symbol).strip() == "":
                continue
            h = {
                "symbol": str(symbol).strip(),
                "name": str(row[1] or symbol).strip(),
                "market": str(row[2] or "US").strip().upper(),
                "cost_price": float(row[3] or 0),
                "quantity": int(float(row[4] or 0)),
                "target_weight": float(row[5] or 10) / 100,  # Excel中是百分比数字
                "sector": str(row[6] or "").strip(),
                "notes": str(row[7] or "").strip(),
            }
            # 可选的止损价
            if row[8] and float(row[8]) > 0:
                h["custom_stop_loss"] = float(row[8])
            # 备注
            if row[9]:
                h["notes"] = str(row[9]).strip()
            holdings.append(h)
        print(f"  📊 从 [持仓录入] 读取 {len(holdings)} 条持仓记录")
    else:
        print("  ⚠️ 未找到 [持仓录入] 工作表")

    # ---- 读取关注列表 ----
    watchlist = []
    if "关注列表" in wb.sheetnames:
        ws2 = wb["关注列表"]
        for row in ws2.iter_rows(min_row=4, max_row=50, values_only=True):
            symbol = row[0]
            if not symbol or str(symbol).strip() == "":
                continue
            w = {
                "symbol": str(symbol).strip(),
                "name": str(row[1] or symbol).strip(),
                "market": str(row[2] or "US").strip().upper(),
                "target_entry": float(row[3] or 0),
                "sector": str(row[4] or "").strip(),
                "reason": str(row[5] or "").strip(),
                "priority": str(row[6] or "中").strip(),
            }
            watchlist.append(w)
        print(f"  👁️  从 [关注列表] 读取 {len(watchlist)} 条关注记录")

    # ---- 读取风险参数 ----
    risk_params = {}
    trading_mode = "稳健"
    base_currency = "CNY"
    if "风险参数" in wb.sheetnames:
        ws3 = wb["风险参数"]
        param_map = {}
        for row in ws3.iter_rows(min_row=4, max_row=15, values_only=True):
            if row[0] and row[1] is not None:
                param_map[str(row[0]).strip()] = row[1]

        trading_mode = str(param_map.get("交易模式", "稳健"))
        base_currency = str(param_map.get("基础货币", "CNY"))

        risk_params = {
            "max_single_position": float(param_map.get("最大单票仓位", 15)) / 100,
            "max_total_position": float(param_map.get("总仓位上限", 80)) / 100,
            "stop_loss_pct": float(param_map.get("止损幅度", -8)) / 100,
            "take_profit_pct": float(param_map.get("止盈目标", 30)) / 100,
            "rebalance_threshold": float(param_map.get("再平衡阈值", 5)) / 100,
            "max_correlated_exposure": float(param_map.get("最大关联暴露", 40)) / 100,
            "max_holdings": int(float(param_map.get("最大持仓数", 15))),
        }
        print(f"  ⚙️  从 [风险参数] 读取交易模式: {trading_mode}")

    wb.close()
    return holdings, watchlist, risk_params, trading_mode, base_currency


# ============================================================
# API导入
# ============================================================

def import_from_api(api_url=None, api_data=None):
    """
    从API接口导入持仓数据
    支持的JSON格式:
    {
      "holdings": [
        {"symbol": "AAPL", "name": "苹果", "market": "US", "cost_price": 165, "quantity": 100, ...},
        ...
      ],
      "watchlist": [...],   // 可选
      "risk_params": {...}  // 可选
    }
    """
    if api_data:
        data = api_data
    elif api_url:
        from urllib.request import urlopen, Request
        req = Request(api_url, headers={"User-Agent": "TradingAssistant/1.0"})
        with urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    else:
        print("❌ 未提供API URL或数据")
        return [], [], {}, "稳健", "CNY"

    holdings = data.get("holdings", [])
    watchlist = data.get("watchlist", [])
    risk_params = data.get("risk_params", {})
    mode = data.get("trading_mode", "稳健")
    currency = data.get("base_currency", "CNY")

    # 标准化字段
    for h in holdings:
        h.setdefault("market", "US")
        h.setdefault("target_weight", 0.10)
        h.setdefault("sector", "")
        h.setdefault("notes", "API导入")
        if isinstance(h.get("target_weight"), (int, float)) and h["target_weight"] > 1:
            h["target_weight"] = h["target_weight"] / 100

    for w in watchlist:
        w.setdefault("market", "US")
        w.setdefault("target_entry", 0)
        w.setdefault("reason", "API导入")

    print(f"  📊 从API读取 {len(holdings)} 条持仓, {len(watchlist)} 条关注")
    return holdings, watchlist, risk_params, mode, currency


# ============================================================
# 合并与写入
# ============================================================

def merge_portfolios(existing_holdings, new_holdings):
    """合并持仓：以新数据为准，symbol相同时用新数据覆盖"""
    merged = {}
    for h in existing_holdings:
        merged[h["symbol"]] = h
    for h in new_holdings:
        merged[h["symbol"]] = h  # 新数据覆盖
    return list(merged.values())


def write_config(holdings, watchlist, risk_params, mode, currency, api_keys=None):
    """写入portfolio-config.json"""
    # 读取现有配置以保留api_keys
    existing = {}
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)

    existing_keys = existing.get("api_keys", {})
    if api_keys:
        existing_keys.update(api_keys)

    config = {
        "meta": {
            "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "trading_mode": mode,
            "base_currency": currency,
            "data_source": "excel+api",
        },
        "risk_params": risk_params if risk_params else existing.get("risk_params", {
            "max_single_position": 0.15,
            "max_total_position": 0.80,
            "stop_loss_pct": -0.08,
            "take_profit_pct": 0.30,
            "max_correlated_exposure": 0.40,
            "rebalance_threshold": 0.05,
            "max_holdings": 15,
        }),
        "portfolio": {
            "holdings": holdings,
            "watchlist": watchlist,
        },
        "api_keys": existing_keys,
    }

    with open(CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 配置已写入: {CONFIG_PATH}")
    print(f"   持仓: {len(holdings)}只 | 关注: {len(watchlist)}只 | 模式: {mode}")


# ============================================================
# 主入口
# ============================================================

def main():
    print("=" * 50)
    print("📥 持仓导入工具")
    print("=" * 50)

    args = sys.argv[1:]

    if "--api" in args:
        # API模式
        idx = args.index("--api")
        api_url = args[idx + 1] if idx + 1 < len(args) else None
        if not api_url:
            print("❌ 用法: python import_portfolio.py --api <url>")
            sys.exit(1)
        print(f"\n🔗 从API导入: {api_url}")
        holdings, watchlist, risk_params, mode, currency = import_from_api(api_url=api_url)
        write_config(holdings, watchlist, risk_params, mode, currency)

    elif "--merge" in args:
        # 合并模式: 先读Excel，再读API
        xlsx_path = DEFAULT_XLSX
        for a in args:
            if a.endswith(".xlsx"):
                xlsx_path = a
                break

        print(f"\n📑 从Excel读取: {xlsx_path}")
        h1, w1, rp, mode, currency = import_from_excel(xlsx_path)

        api_url = None
        if "--api-url" in args:
            idx = args.index("--api-url")
            api_url = args[idx + 1] if idx + 1 < len(args) else None

        if api_url:
            print(f"\n🔗 从API补充: {api_url}")
            h2, w2, rp2, _, _ = import_from_api(api_url=api_url)
            h1 = merge_portfolios(h1, h2)
            w1 = merge_portfolios(w1, w2)
            if rp2:
                rp.update(rp2)

        write_config(h1, w1, rp, mode, currency)

    elif "--from-json" in args:
        # 从JSON文件导入（方便程序化调用）
        idx = args.index("--from-json")
        json_path = args[idx + 1] if idx + 1 < len(args) else None
        if not json_path or not os.path.exists(json_path):
            print("❌ 用法: python import_portfolio.py --from-json <path.json>")
            sys.exit(1)
        print(f"\n📄 从JSON导入: {json_path}")
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        holdings, watchlist, risk_params, mode, currency = import_from_api(api_data=data)
        write_config(holdings, watchlist, risk_params, mode, currency)

    else:
        # 默认: 从Excel导入
        xlsx_path = args[0] if args and args[0].endswith(".xlsx") else DEFAULT_XLSX
        print(f"\n📑 从Excel导入: {xlsx_path}")
        holdings, watchlist, risk_params, mode, currency = import_from_excel(xlsx_path)
        write_config(holdings, watchlist, risk_params, mode, currency)


if __name__ == "__main__":
    main()
